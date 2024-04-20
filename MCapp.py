### Developed by Andrei Costan with Python 3.9.7

import sys, os, re, json
from PyQt5.uic import loadUi
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import QDialog, QApplication, QStackedWidget, QPushButton, QLabel, QComboBox, QMessageBox, QTableWidget, QTableWidgetItem, QLineEdit, QPlainTextEdit, QTabWidget, QDateEdit
from datetime import date, datetime
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np

### GUI
class UserWelcomeWindow(QDialog):
    def __init__(self):
        super(UserWelcomeWindow, self).__init__()
        loadUi('UIs\\UserWelcomeWindow-MCapp.ui', self)

        global weeks_T1, number_of_weeks_T1

        self.to_email_sms_custom_button = self.findChild(QPushButton, 'to_email_sms_custom_button')
        self.to_email_sms_custom_button.clicked.connect(lambda: self.go_to_EmailSMSWindow(number_of_weeks_T1, 'trig-custom'))
        self.to_email_sms_3weeks_button = self.findChild(QPushButton, 'to_email_sms_3weeks_button')
        self.to_email_sms_3weeks_button.clicked.connect(lambda: self.go_to_EmailSMSWindow(3, 'trig-3'))
        self.to_email_sms_5weeks_button = self.findChild(QPushButton, 'to_email_sms_5weeks_button')
        self.to_email_sms_5weeks_button.clicked.connect(lambda: self.go_to_EmailSMSWindow(5, 'trig-5'))
        self.to_telephone_list_2weeks = self.findChild(QPushButton, 'to_telephone_list_2weeks')
        self.to_telephone_list_2weeks.clicked.connect(lambda: self.go_to_EmailSMSWindow(2, 'trig-2'))
        self.to_add_in_database_button = self.findChild(QPushButton, 'to_add_in_database_button')
        self.to_add_in_database_button.clicked.connect(self.go_to_AddInDatabase)
        self.combo_weeks_T1 = self.findChild(QComboBox, 'combo_weeks_T1')
        self.combo_weeks_T1.addItems(weeks_T1)
        self.combo_weeks_T1.setCurrentIndex(2)
        self.combo_weeks_T1.activated.connect(self.get_number_of_weeks_T1)

    def get_number_of_weeks_T1(self):
        global number_of_weeks_T1
        number_of_weeks_T1 = int(self.combo_weeks_T1.currentText())

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Escape:
            pass

    def go_to_EmailSMSWindow(self, number_of_weeks, button_trigger):
        global number_of_weeks_T1, custom_title_label
        number_of_weeks_T1 = number_of_weeks
        ### Title Label Set Text
        if button_trigger == 'trig-2':
            custom_title_label = 'MCapp\nLista Telefoane 2 Saptamani'
        elif button_trigger == 'trig-3':
            custom_title_label = 'MCapp\nTrimiteri Email / SMS 3 Saptamani'
        elif button_trigger == 'trig-5':
            custom_title_label = 'MCapp\nTrimiteri Email / SMS 5 Saptamani'
        elif button_trigger == 'trig-custom':
            custom_title_label = f'MCapp - Trimiteri Email / SMS\nSelectie Manuala: {number_of_weeks_T1} Saptamani'
        else:
            custom_title_label = 'MCapp\nTrimiteri Email / SMS'

        job_select_page = EmailSMSWindow()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def go_to_AddInDatabase(self):
        job_select_page = AddInDatabase()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)

class EmailSMSWindow(QDialog):
    def __init__(self):
        super(EmailSMSWindow, self).__init__()
        loadUi('UIs\\EmailSMSWindow.ui', self)

        global link_firma_programari, telefon_programari, email_firma, data_azi, column_name_list, df_filtered, df_filtered_small, number_of_weeks_T1, custom_title_label
        
        load_DB_for_mailings()

        self.label_title = self.findChild(QLabel, 'label_title')
        self.label_title.setText(custom_title_label)

        self.back_user_welcome_window = self.findChild(QPushButton, 'back_user_welcome_window')
        self.back_user_welcome_window.clicked.connect(self.go_to_UserWelcomeWindow)
        self.refresh_database_button = self.findChild(QPushButton, 'refresh_database_button')
        self.refresh_database_button.clicked.connect(self.reload_mailing_database)

        self.table = self.findChild(QTableWidget, 'tableWidget_1')
        self.table.setColumnCount(len(column_name_list)+1)
        self.table.setHorizontalHeaderLabels(['']+column_name_list)

        self.message_dict = df_filtered.to_dict('records')
        
        for index, row in df_filtered_small.iterrows():
            self.add_table_row(self.table, ['']+row.to_list())
        
        for index in range(self.table.rowCount()):
            self.btn = QPushButton(self.table)
            self.btn.setText('open: ' + str(index+1))
            self.table.setCellWidget(index, 0, self.btn)
            self.btn.clicked.connect(lambda state, index=index:self.clickerTable(index, self.message_dict))
        
        self.table.resizeColumnsToContents()

    def clickerTable(self, index, message_dict):
        global number_of_weeks_T1, custom_title_label
        self.dlg = MessageWindow()
        self.tabs = self.dlg.findChild(QTabWidget, 'tabWidget')
        ### Buttons
        self.sms_telephone_number_button = self.dlg.findChild(QPushButton, 'sms_telephone_number_button')
        self.sms_message_button = self.dlg.findChild(QPushButton, 'sms_message_button')
        self.add_sms_sent_button = self.dlg.findChild(QPushButton, 'add_sms_sent_button')
        self.email_to_button = self.dlg.findChild(QPushButton, 'email_to_button')
        self.email_bcc_button = self.dlg.findChild(QPushButton, 'email_bcc_button')
        self.email_subject_button = self.dlg.findChild(QPushButton, 'email_subject_button')
        self.email_message_button = self.dlg.findChild(QPushButton, 'email_message_button')
        self.add_email_sent_button = self.dlg.findChild(QPushButton, 'add_email_sent_button')
        ### Labels and Texts
        self.nume_client_SMS = self.dlg.findChild(QLabel, 'nume_client_var')
        self.telephone_text_SMS = self.dlg.findChild(QPlainTextEdit, 'telephone_text')
        self.body_SMS = self.dlg.findChild(QPlainTextEdit, 'sms_body')
        self.nume_client_email = self.dlg.findChild(QLabel, 'nume_client_var_2')
        self.email_client = self.dlg.findChild(QPlainTextEdit, 'email_client')
        self.email_contact = self.dlg.findChild(QPlainTextEdit, 'email_contact')
        self.subject_email = self.dlg.findChild(QPlainTextEdit, 'subject_text')
        self.body_email = self.dlg.findChild(QPlainTextEdit, 'email_body')        
        
        ### Remove certain Tabs
        if message_dict[index]['Telefon'] == '-' and message_dict[index]['Email'] == '-':
            self.tabs.removeTab(1)
            self.tabs.removeTab(0)
        if message_dict[index]['Telefon'] == '-':
            self.tabs.removeTab(0)
        if message_dict[index]['Email'] == '-':
            self.tabs.removeTab(1)
        
        self.client_name = message_dict[index]['Persoana contact']
        self.sms_telephone_number = message_dict[index]['Telefon'].replace('.', '')
        self.sms_message = SMS_body_generator(self.client_name, message_dict[index]['Data expirare'], message_dict[index]['Adresa client'], telefon_programari, link_firma_programari)
        self.email_to = message_dict[index]['Email']
        self.email_bcc = email_firma
        self.email_subject = f'Expirare ceva'
        self.email_message = email_body_generator(self.client_name, message_dict[index]['Data expirare'], message_dict[index]['Adresa client'], telefon_programari, link_firma_programari)
        ### Set Texts
        self.nume_client_SMS.setText(f'{self.client_name}')
        self.telephone_text_SMS.setPlainText(f'{self.sms_telephone_number}')
        self.body_SMS.setPlainText(f'{self.sms_message}')
        self.nume_client_email.setText(f'{self.client_name}')
        self.email_client.setPlainText(f'{self.email_to}')
        self.email_contact.setPlainText(f'{email_firma}')
        self.subject_email.setPlainText(f'{self.email_subject}')
        self.body_email.setPlainText(self.email_message)
        
        self.sms_telephone_number_button.clicked.connect(lambda: self.copy_text_to_clipboard(self.sms_telephone_number))
        self.sms_message_button.clicked.connect(lambda: self.copy_text_to_clipboard(self.sms_message))
        self.email_to_button.clicked.connect(lambda: self.copy_text_to_clipboard(self.email_to))
        self.email_bcc_button.clicked.connect(lambda: self.copy_text_to_clipboard(self.email_bcc))
        self.email_subject_button.clicked.connect(lambda: self.copy_text_to_clipboard(self.email_subject))
        self.email_message_button.clicked.connect(lambda: self.copy_text_to_clipboard(self.email_message))
        
        if custom_title_label == 'MCapp\nLista Telefoane 2 Saptamani':
            self.add_sms_sent_button.clicked.connect(lambda: self.excel_cell_write(message_dict[index]['ID'], 'telephone list'))
        elif number_of_weeks_T1 >= 4:
            self.add_sms_sent_button.clicked.connect(lambda: self.excel_cell_write(message_dict[index]['ID'], 'sms t2'))
            self.add_email_sent_button.clicked.connect(lambda: self.excel_cell_write(message_dict[index]['ID'], 'email t2'))
        else:
            self.add_sms_sent_button.clicked.connect(lambda: self.excel_cell_write(message_dict[index]['ID'], 'sms t1'))
            self.add_email_sent_button.clicked.connect(lambda: self.excel_cell_write(message_dict[index]['ID'], 'email t1'))
        
        self.dlg.open()

    def reload_mailing_database(self):
        job_select_page = UserWelcomeWindow()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)
        job_select_page = EmailSMSWindow()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)
        refreshed_popup()

    def copy_text_to_clipboard(self, text):
        cb = QApplication.clipboard()
        cb.clear(mode=cb.Clipboard)
        cb.setText(text, mode=cb.Clipboard)

    def excel_cell_write(self, idx, col_type):
        if col_type == 'sms t1':
            col = 24
        elif col_type == 'email t1':
            col = 25
        elif col_type == 'sms t2':
            col = 26
        elif col_type == 'email t2':
            col = 27
        elif col_type == 'telephone list':
            col = 23
        try:
            wb = xl.load_workbook(database_location)
            ws = wb['Baza de Date']
            ws.cell(row=idx+1, column=col, value=f'DA {data_azi.strftime("%d.%m.%Y")}')
            wb.save(database_location)
            self.added_info()
        except:
            close_excel_popup()

    def added_info(self):
        reply = QMessageBox.question(None, 'Adaugat Informatie in Baza de Date', 'Informatiile au fost adaugate in Baza de date\nContinuati?',
        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

        if reply == QMessageBox.No:
            self.dlg.reject()

    def add_table_row(self, table, rowData):
        row = table.rowCount()
        table.setRowCount(row+1)
        col = 0
        for item in rowData:
            cell = QTableWidgetItem(str(item))
            table.setItem(row, col, cell)
            col += 1

    def go_to_UserWelcomeWindow(self):
        job_select_page = UserWelcomeWindow()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)
        global number_of_weeks_T1
        number_of_weeks_T1 = 3

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Escape:
            pass

class MessageWindow(QDialog):
    def __init__(self):
        super(MessageWindow, self).__init__()
        loadUi('UIs\\ClientMessages.ui', self)

class AddInDatabase(QDialog):
    def __init__(self):
        super(AddInDatabase, self).__init__()
        loadUi('UIs\\AddInDatabase.ui', self)
        self.initial_values_allocation()
        ### Back to UserWelcomeWindow Button
        self.back_user_welcome_window = self.findChild(QPushButton, 'back_user_welcome_window')
        self.back_user_welcome_window.clicked.connect(self.go_to_UserWelcomeWindow)
        ### add_record_in_database_button
        self.add_record_in_database_button = self.findChild(QPushButton, 'add_record_in_database_button')
        self.add_record_in_database_button.clicked.connect(self.excel_write_new_excel_row)
        ### QDateEdits
        self.dateEdit_data_incheiere_contract = self.findChild(QDateEdit, 'dateEdit_data_incheiere_contract')
        self.dateEdit_data_incheiere_contract.setDate(get_today())
        self.dateEdit_data_finalizare_contract = self.findChild(QDateEdit, 'dateEdit_data_finalizare_contract')
        self.dateEdit_data_finalizare_contract.setDate(get_today())
        self.dateEdit_data_inregistrare_distrigaz = self.findChild(QDateEdit, 'dateEdit_data_inregistrare_distrigaz')
        self.dateEdit_data_inregistrare_distrigaz.setDate(get_today())
        ### QLineEdits
        self.lineEdit_id = self.findChild(QLineEdit, 'lineEdit_id')
        self.lineEdit_id.setText(f'{self.id_initial}')
        self.lineEdit_anul = self.findChild(QLineEdit, 'lineEdit_anul')
        self.lineEdit_anul.setText(f'{self.an_curent}')
        self.lineEdit_id_client = self.findChild(QLineEdit, 'lineEdit_id_client')
        self.lineEdit_numar_contract = self.findChild(QLineEdit, 'lineEdit_numar_contract')
        self.lineEdit_persoana_contact = self.findChild(QLineEdit, 'lineEdit_persoana_contact')
        self.lineEdit_telefon = self.findChild(QLineEdit, 'lineEdit_telefon')
        self.lineEdit_email = self.findChild(QLineEdit, 'lineEdit_email')
        self.lineEdit_numar_inregistrare_distrigaz = self.findChild(QLineEdit, 'lineEdit_numar_inregistrare_distrigaz')
        self.lineEdit_puncte_de_ardere = self.findChild(QLineEdit, 'lineEdit_puncte_de_ardere')
        self.lineEdit_kit_detectie_gaze = self.findChild(QLineEdit, 'lineEdit_kit_detectie_gaze')
        self.lineEdit_apreciere_contract_finalizat = self.findChild(QLineEdit, 'lineEdit_apreciere_contract_finalizat')
        self.lineEdit_apreciere_contract_finalizat.setText(f'10')
        self.lineEdit_numar_strada = self.findChild(QLineEdit, 'lineEdit_numar_strada')
        self.lineEdit_bloc = self.findChild(QLineEdit, 'lineEdit_bloc')
        self.lineEdit_scara = self.findChild(QLineEdit, 'lineEdit_scara')
        self.lineEdit_etaj = self.findChild(QLineEdit, 'lineEdit_etaj')
        self.lineEdit_apartament = self.findChild(QLineEdit, 'lineEdit_apartament')
        self.lineEdit_alte_detalii = self.findChild(QLineEdit, 'lineEdit_alte_detalii')
        ### QComboBoxes
        self.comboBox_judet = self.findChild(QComboBox, 'comboBox_judet')
        self.comboBox_oras = self.findChild(QComboBox, 'comboBox_oras')
        for k, v in judet_sat_dict.items():
            self.comboBox_judet.addItem(k, v)
        self.comboBox_judet.activated.connect(self.clicker_combo_judet)
        self.comboBox_oras.activated.connect(self.clicker_combo_oras)
        self.comboBox_tip_artera = self.findChild(QComboBox, 'comboBox_tip_artera')
        self.comboBox_tip_artera.addItems(tip_artera)
        self.comboBox_tip_artera.activated.connect(self.clicker_combo_artera)
        self.comboBox_strada = self.findChild(QComboBox, 'comboBox_strada')
        self.comboBox_strada.setPlaceholderText('...')
        self.comboBox_strada.activated.connect(self.clicker_combo_strada)
        self.comboBox_tip_proprietar = self.findChild(QComboBox, 'comboBox_tip_proprietar')
        self.comboBox_tip_proprietar.addItems(tip_proprietar)
        self.comboBox_tip_verificare_revizie = self.findChild(QComboBox, 'comboBox_tip_verificare_revizie')
        self.comboBox_tip_verificare_revizie.addItems(tip_verificare_revizie)
        self.comboBox_instalator_nume_prenume = self.findChild(QComboBox, 'comboBox_instalator_nume_prenume')
        for k, v in nume_instalator.items():
            self.comboBox_instalator_nume_prenume.addItem(k, v)
        self.comboBox_instalator_cnp_nr_leg = self.findChild(QComboBox, 'comboBox_instalator_cnp_nr_leg')
        self.comboBox_instalator_tip_leg = self.findChild(QComboBox, 'comboBox_instalator_tip_leg')
        self.comboBox_instalator_nume_prenume.activated.connect(self.clicker_combo_instalator)
        self.comboBox_instalator_cnp_nr_leg.activated.connect(self.clicker_combo_instalator_cnp_nrleg)
        self.comboBox_instalator_tip_leg.activated.connect(self.clicker_combo_instalator_legitimatie)

    def initial_values_allocation(self):
        ### Initial variables values allocation
        self.id_initial = self.get_id_from_database()
        self.an_curent = get_today().year
        self.numar_contract = '-'
        self.id_client = '-'
        self.judet = '-'
        self.oras = '-'
        self.artera = 'Strada'
        self.strada = '-'
        self.numar_strada = '-'
        self.bloc = '-'
        self.scara = '-'
        self.etaj = '-'
        self.apartament = '-'
        self.alte_detalii = '-'
        self.tip_proprietar = 'Fizica'
        self.persoana_contact = '-'
        self.telefon = '-'
        self.email = '-'
        self.numar_inregistrare_distrigaz = '-'
        self.kit_detectie_gaze = '-'
        self.puncte_de_ardere = '-'
        self.tip_verificare_revizie = '-'
        self.instalator_nume_prenume = '-'
        self.instalator_cnp_nr_leg = '-'
        self.instalator_tip_leg = '-'
        self.apreciere_contract_finalizat = '10'

    def completed_values_allocation(self):
        ### Completed variables values allocation
        self.id = self.lineEdit_id.text()
        self.anul = self.lineEdit_anul.text()
        self.numar_contract = self.lineEdit_numar_contract.text()
        self.id_client = self.lineEdit_id_client.text()
        self.judet = self.comboBox_judet.currentText()
        self.oras = self.comboBox_oras.currentText()
        self.artera = self.comboBox_tip_artera.currentText()
        self.strada = self.comboBox_strada.currentText()
        self.numar_strada = self.lineEdit_numar_strada.text()
        self.bloc = self.lineEdit_bloc.text()
        self.scara = self.lineEdit_scara.text()
        self.etaj = self.lineEdit_etaj.text()
        self.apartament = self.lineEdit_apartament.text()
        self.alte_detalii = self.lineEdit_alte_detalii.text()
        self.tip_proprietar = self.comboBox_tip_proprietar.currentText()
        self.persoana_contact = self.lineEdit_persoana_contact.text()
        self.telefon = self.lineEdit_telefon.text()
        self.email = self.lineEdit_email.text()
        self.numar_inregistrare_distrigaz = self.lineEdit_numar_inregistrare_distrigaz.text()
        self.kit_detectie_gaze = self.lineEdit_kit_detectie_gaze.text()
        self.puncte_de_ardere = self.lineEdit_puncte_de_ardere.text()
        self.tip_verificare_revizie = self.comboBox_tip_verificare_revizie.currentText()
        self.instalator_nume_prenume = self.comboBox_instalator_nume_prenume.currentText()
        self.instalator_cnp_nr_leg = self.comboBox_instalator_cnp_nr_leg.currentText()
        self.instalator_tip_leg = self.comboBox_instalator_tip_leg.currentText()
        self.apreciere_contract_finalizat = self.lineEdit_apreciere_contract_finalizat.text()

    def get_id_from_database(self):
        self.df_baza_add = pd.read_excel(database_location, sheet_name='Baza de Date', dtype='object', keep_default_na=False, engine='openpyxl')
        self.new_id = str(self.df_baza_add.ID.size+1)
        return self.new_id

    def clicker_combo_instalator(self, index):
        self.comboBox_instalator_cnp_nr_leg.clear()
        self.comboBox_instalator_tip_leg.clear()
        self.comboBox_instalator_cnp_nr_leg.addItem(self.comboBox_instalator_nume_prenume.itemData(index)[0])
        self.comboBox_instalator_tip_leg.addItem(self.comboBox_instalator_nume_prenume.itemData(index)[1])
        self.comboText_instalator_nume_prenume = self.comboBox_instalator_nume_prenume.currentText()
        self.comboText_instalator_cnp_nr_leg = self.comboBox_instalator_cnp_nr_leg.currentText()
        self.comboText_instalator_tip_leg = self.comboBox_instalator_tip_leg.currentText()

    def clicker_combo_instalator_cnp_nrleg(self):
        self.comboText_instalator_cnp_nr_leg = self.comboBox_instalator_cnp_nr_leg.currentText()

    def clicker_combo_instalator_legitimatie(self):
        self.comboText_instalator_tip_leg = self.comboBox_instalator_tip_leg.currentText()

    def clicker_combo_judet(self, index):
        self.comboBox_oras.clear()
        self.comboBox_tip_artera.clear()
        self.comboBox_tip_artera.addItems(tip_artera)
        self.comboBox_oras.addItems(self.comboBox_judet.itemData(index))
        self.comboText_judet = self.comboBox_judet.currentText()
        self.comboText_oras = self.comboBox_oras.currentText()
        self.combo_add_str_art()

    def combo_add_str_art(self):
        if self.comboText_judet in streets_dict:
            if self.comboText_oras in streets_dict.get(self.comboText_judet):
                self.comboBox_strada.clear()
                for k, v in streets_dict.get(self.comboText_judet).get(self.comboText_oras).items():
                    self.comboBox_strada.addItem(k, v)
            else:
                self.comboBox_tip_artera.clear()
                self.comboBox_tip_artera.addItems(tip_artera)
                self.comboBox_strada.clear()
        else:
            self.comboBox_strada.clear()

    def clicker_combo_oras(self):
        self.comboBox_strada.clear()
        self.comboBox_tip_artera.clear()
        self.comboBox_tip_artera.addItems(tip_artera)
        self.comboText_oras = self.comboBox_oras.currentText()
        self.combo_add_str_art()

    def clicker_combo_strada(self, index):
        self.comboBox_tip_artera.clear()
        self.comboBox_tip_artera.addItems(self.comboBox_strada.itemData(index))
        self.comboText_strada = self.comboBox_strada.currentText()

    def clicker_combo_artera(self):
        self.comboText_artera = self.comboBox_tip_artera.currentText()

    def excel_write_new_excel_row(self):
        ### Verifs False
        self.verif_anul = False
        self.verif_id = False
        self.verif_id_client = False
        self.verif_judet = False
        self.verif_instalator = False
        self.verif_persoana_contact = False
        self.verif_numar_contract = False
        self.verif_numar_inregistrare_distrigaz = False
        self.verif_email = False
        self.verif_telefon = False
        self.verif_data_incheiere_finalizare_contract = False
        self.verif_strada = False
        self.df_baza_add = pd.concat([self.df_baza_add, self.get_database_excel_row()], ignore_index=True)
        ### Verifications List
        self.verif_list = [
            self.verif_anul,
            self.verif_id,
            self.verif_judet,
            self.verif_instalator,
            self.verif_id_client,
            self.verif_persoana_contact,
            self.verif_numar_contract,
            self.verif_numar_inregistrare_distrigaz,
            self.verif_email,
            self.verif_telefon,
            self.verif_data_incheiere_finalizare_contract,
            self.verif_strada]
        if False not in self.verif_list:
            try:
                ### Excel Open and Write
                wb = xl.load_workbook(database_location)
                ws = wb['Baza de Date']
                for r in dataframe_to_rows(self.get_database_excel_row(), index=False, header=False):
                    ws.append(r)
                for i in range(1, 28):
                    ws.cell(row=int(self.id)+1, column=i)._style = ws.cell(row=int(self.id), column=i)._style
                wb.save(database_location)
                self.info_popup(f'Informatiile au fost adaugate in Baza de Date!\nApasati OK si asteptati pana se reseteaza pagina!')
                self.reload_AddInDatabase_window()
            except:
                close_excel_popup()

    def get_database_excel_row(self):
        self.completed_values_allocation()
        self.empty_values_verification()
        self.new_row_dict = {
            'ID': [int(self.id)],
            'Anul': [int(self.anul)],
            'Numar contract': [self.numar_contract],
            'Data incheierii contractului': [self.dateEdit_data_incheiere_contract.date().toPyDate().strftime('%d.%m.%Y')],
            'Data finalizarii contractului': [self.dateEdit_data_finalizare_contract.date().toPyDate().strftime('%d.%m.%Y')],
            'Tip proprietar': [self.tip_proprietar],
            'Adresa client': [self.get_address_concat()],
            'Oras': [self.oras],
            'Judet': [self.judet],
            'ID Client': [self.id_client],
            'Persoana contact': [re.sub(r'\s+', ' ', self.persoana_contact).title().strip()],
            'Telefon': [self.telefon],
            'Email': [re.sub(r'\s+', '', self.email).strip()],
            'Numar inregistrare Distrigaz': [self.numar_inregistrare_distrigaz],
            'Data inregistrare Distrigaz': [self.dateEdit_data_inregistrare_distrigaz.date().toPyDate().strftime('%d.%m.%Y')],
            'Puncte de ardere': [self.puncte_de_ardere],
            'Kit detectie gaze': [self.kit_detectie_gaze],
            'Tip verificare / revizie': [self.tip_verificare_revizie],
            'Instalatorul / instalatorii autorizat(ii) ANRE care au efectuat lucrarea - nume, prenume': [self.instalator_nume_prenume],
            'Instalatorul / instalatorii autorizat(ii) ANRE care au efectuat lucrarea - CNP / Nr Legitimatie': [self.instalator_cnp_nr_leg],
            'Instalatorul / instalatorii autorizat(ii) ANRE care au efectuat lucrarea - tipul legitimatiei': [self.instalator_tip_leg],
            'Apreciere contract finalizat (valoarea gradului de satisfactie)': [int(self.apreciere_contract_finalizat)],
            'PROGRAMAT DA/NU': [''],
            'SMS 3S DA/NU': [''],
            'EMAIL 3S DA/NU': [''],
            'SMS 5S DA/NU': [''],
            'EMAIL 5S DA/NU': ['']
        }
        self.df_new_row = pd.DataFrame.from_dict(self.new_row_dict, dtype=object)
        return self.df_new_row

    def get_address_concat(self):
        ### Address Building Function
        if not re.search(r'^(\-)$', self.strada, re.I):
            self.strada = re.sub(r'\s+', ' ', self.strada).title().strip()
        if not re.search(r'^(\-)$', self.numar_strada, re.I):
            self.numar_strada = re.sub(r'(\,|\.|\;|\s+)', ' ', self.numar_strada).strip()
        if not re.search(r'^(\-)$', self.bloc, re.I):
            self.bloc = re.sub(r'(\,|\.|\;|\s+)', ' ', self.bloc).strip()
        if not re.search(r'^(\-)$', self.scara, re.I):
            self.scara = re.sub(r'(\,|\.|\;|\s+)', ' ', self.scara).strip()
        if not re.search(r'^(\-)$', self.etaj, re.I):
            self.etaj = re.sub(r'(\,|\.|\;|\s+)', ' ', self.etaj).strip()
        if not re.search(r'^(\-)$', self.apartament, re.I):
            self.apartament = re.sub(r'(\,|\.|\;|\s+)', ' ', self.apartament).strip()
        if not re.search(r'^(\-)$', self.alte_detalii, re.I):
            self.alte_detalii = re.sub(r'(\s+)', ' ', self.alte_detalii).strip()
        
        self.full_adress = (
            f'{self.artera} '
            f'{self.strada}, '
            f'Nr. {self.numar_strada}, '
            f'Bl. {self.bloc}, '
            f'Sc. {self.scara}, '
            f'Et. {self.etaj}, '
            f'Ap. {self.apartament}, '
            f'{self.alte_detalii}, '
            f'{self.oras}, '
            f'{self.judet}'
        )
        return self.full_adress

    def empty_values_verification(self):
        # id Verifs
        if self.id == '' or re.search(r'^\s+$', self.id, re.I) or str(self.id) == '0' or self.id == None:
            self.warning_popup('id', f'Idul este gol\nAlocam id_initial la Id?')
        elif int(self.id_initial) > int(self.id):
            self.warning_popup('id', f'Idul este deja existent in baza de date!\nSe va atribui id_initial la Id!')
        elif int(self.id_initial) == int(self.id):
            self.verif_id = True
        else:
            self.warning_popup('id', f'ID modificat si necontinuu!\nAlocam id_initial la Id?')

        # anul Verifs
        if self.anul == '' or re.search(r'^\s+$', self.anul, re.I) or str(self.anul) == '0' or self.anul == None:
            self.warning_popup('anul', f'Anul este gol\nAlocam an_curent la an?')
        elif str(self.an_curent) != str(self.anul):
            self.verif_anul = True
        else:
            self.verif_anul = True

        # id client Verifs
        if self.id_client == '' or self.id_client == '-' or re.search(r'^\s+$', self.id_client, re.I) or self.id_client == None:
            self.info_popup(f'Nu ai completat ID Client!\nID Client trebuie completat obligatoriu!')
        else:
            self.verif_id_client = True

        # judetul Verifs
        if self.judet == '...Alege un Judet...' or self.judet == '' or self.judet == None:
            self.info_popup(f'Nu ai ales un Judet!\nAlege un judet pentru a putea continua!')
        else:
            self.verif_judet = True

        # instalator Verifs
        if self.instalator_nume_prenume == '...Alege Nume' or self.instalator_nume_prenume == '' or self.instalator_nume_prenume == None:
            self.info_popup(f'Nu ai ales un Instalator!\nAlege un instalator pentru a putea continua!')
        else:
            self.verif_instalator = True

        # persoana contact Verifs
        if self.persoana_contact == '' or self.persoana_contact == '-' or re.search(r'^\s+$', self.persoana_contact, re.I) or self.persoana_contact == None:
            self.info_popup(f'Nu ai completat Persoana Contact!\nPersoana Contact trebuie completata obligatoriu!')
        else:
            self.verif_persoana_contact = True

        # numar de inregistrare Distrigaz Verifs
        if self.numar_inregistrare_distrigaz == '' or re.search(r'^\s+$', self.numar_inregistrare_distrigaz, re.I) or self.numar_inregistrare_distrigaz == None:
            self.warning_popup('numar_inregistrare_distrigaz', f'Nu ai completat Numar Inregistrare Distrigaz!\nContinuati cu Numar Inregistrare Distrigaz NECOMPLETAT?')
        else:
            self.verif_numar_inregistrare_distrigaz = True

        # numar contract Verifs
        if self.numar_contract == '' or re.search(r'^\s+$', self.numar_contract, re.I) or self.numar_contract == None:
            self.warning_popup('numar_contract', f'Nu ai completat Numar Contract!\nContinuati cu Numar Contract NECOMPLETAT?')
        else:
            self.verif_numar_contract = True

        # telefon Verifs
        if self.telefon == '' or re.search(r'^\s+$', self.telefon, re.I) or self.telefon == None:
            self.warning_popup('telefon', f'Nu ai completat Telefon!\nContinuati cu Telefon NECOMPLETAT?')
        else:
            self.verif_telefon = True

        # email Verifs
        if self.email == '' or re.search(r'^\s+$', self.email, re.I) or self.email == None:
            self.warning_popup('email', f'Nu ai completat Email!\nContinuati cu Email NECOMPLETAT?')
        elif self.email != '-':
            if re.search(r'^.+@.+\..+$', self.email, re.I):
                self.verif_email = True
            else:
                self.info_popup(f'Email probabil nu este scris corespunzator! (ex: abc@gmail.com)\nVerificati si corectati!')
        else:
            self.verif_email = True

        # incheiere > finalizare Verifs
        if self.dateEdit_data_incheiere_contract.date().toPyDate() > self.dateEdit_data_finalizare_contract.date().toPyDate():
            self.info_popup(f'Data Finalizare Contract este anterioara datei Incheiere Contract!\nVerificati si corectati!')
        else:
            self.verif_data_incheiere_finalizare_contract = True

        # strada - la sectoare
        if self.oras in ['Sector 1', 'Sector 2', 'Sector 3', 'Sector 4', 'Sector 5', 'Sector 6'] and (self.strada == '' or self.strada == '-' or re.search(r'^\s+$', self.strada, re.I) or self.strada == None):
            self.info_popup(f'Strada trebuie selecatata sau completata la Bucuresti!\nVerificati si completati!')
        elif self.strada == '' or re.search(r'^\s+$', self.strada, re.I) or self.strada == None:
            self.strada = '-'
            self.verif_strada = True
        else:
            self.verif_strada = True

        # empty ones
        if self.numar_strada == '' or re.search(r'^\s+$', self.numar_strada, re.I):
            self.numar_strada = '-'
        if self.bloc == '' or re.search(r'^\s+$', self.bloc, re.I):
            self.bloc = '-'
        if self.scara == '' or re.search(r'^\s+$', self.scara, re.I):
            self.scara = '-'
        if self.etaj == '' or re.search(r'^\s+$', self.etaj, re.I):
            self.etaj = '-'
        if self.apartament == '' or re.search(r'^\s+$', self.apartament, re.I):
            self.apartament = '-'
        if self.alte_detalii == '' or re.search(r'^\s+$', self.alte_detalii, re.I):
            self.alte_detalii = '-'
        if self.kit_detectie_gaze == '' or re.search(r'^\s+$', self.kit_detectie_gaze, re.I):
            self.kit_detectie_gaze = '0'
        if self.puncte_de_ardere == '' or re.search(r'^\s+$', self.puncte_de_ardere, re.I):
            self.puncte_de_ardere = '-'
    
    def warning_popup(self, warn_type, txt):
        msg = QMessageBox(QMessageBox.Warning, 'Warning!', f'{txt}', QMessageBox.Cancel|QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Cancel)
        self.warn_type = warn_type
        msg.buttonClicked.connect(self.warning_popup_button)
        msg.exec_()

    def warning_popup_button(self, i):
        # ifs for continue or not / OK | Cancel
        if self.warn_type == 'numar_inregistrare_distrigaz' and i.text() == 'OK':
            self.lineEdit_numar_inregistrare_distrigaz.setText(f'-')
            self.verif_numar_inregistrare_distrigaz = True
        elif self.warn_type == 'numar_contract' and i.text() == 'OK':
            self.lineEdit_numar_contract.setText(f'-')
            self.verif_numar_contract = True
        elif self.warn_type == 'telefon' and i.text() == 'OK':
            self.telefon = '-'
            self.lineEdit_telefon.setText(f'-')
            self.verif_telefon = True
        elif self.warn_type == 'email' and i.text() == 'OK':
            self.email = '-'
            self.lineEdit_email.setText(f'-')
            self.verif_email = True
        elif self.warn_type == 'anul' and i.text() == 'OK':
            self.an_curent = get_today().year
            self.lineEdit_anul.setText(f'{self.an_curent}')
            self.verif_anul = True
        elif self.warn_type == 'id' and i.text() == 'OK':
            self.id_initial = self.get_id_from_database()
            self.lineEdit_id.setText(f'{self.id_initial}')
            self.verif_id = True
        else:
            pass

    def info_popup(self, txt):
        ### Info Popup Warning
        msg = QMessageBox(QMessageBox.Information, 'Warning!', f'{txt}')
        msg.exec_()

    def reload_AddInDatabase_window(self):
        job_select_page = UserWelcomeWindow()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)
        job_select_page = AddInDatabase()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def go_to_UserWelcomeWindow(self):
        job_select_page = UserWelcomeWindow()
        widget.addWidget(job_select_page)
        widget.setCurrentIndex(widget.currentIndex()+1)
        global number_of_weeks_T1
        number_of_weeks_T1 = 3

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Escape:
            pass

def under_construction_popup():
    ### Under Construction Popup Warning
    msg = QMessageBox(QMessageBox.Information, 'UNDER CONSTRUCTION', 'Functionality Is Not Yet Available\nUNDER CONSTRUCTION!!!')
    msg.exec_()

def refreshed_popup():
    ### Refreshed Table Popup Info
    msg = QMessageBox(QMessageBox.Information, 'Table Refreshed', 'Table Refreshed!\nClick OK\nor\npress ESC to continue.')
    msg.exec_()

def close_excel_popup():
    ### Close Excel Popup Warning
    msg = QMessageBox(QMessageBox.Information, 'Inchide Baza de Date Excel', 'Inchide Baza de Date Excel\nNu se poate scrie in Baza cat e deschisa in Excel!')
    msg.exec_()

def SMS_body_generator(nume_contact, data_expirare, adresa_scurta, telefon_programari, site_firma_programari):
    text = f'''Draga Client,
Pe data de: {data_expirare}, va expira ceva, pentru adresa: {adresa_scurta}.
Pentru programari, suna la {telefon_programari} sau online la {site_firma_programari}
Echipa'''
    return text

def email_body_generator(nume_contact, data_expirare, adresa_completa, telefon_programari, site_firma_programari):
    text = f'''Draga Client,

Pe data de: {data_expirare}, va expira ceva, pentru adresa: {adresa_completa}.

Programati-va telefonic la {telefon_programari} sau online la {site_firma_programari}
'''
    return text

def get_today():
    today = date.today()
    return today

def load_DB_for_mailings():
    global row_count, column_name_list, number_of_weeks_T1, data_azi, df_filtered, df_filtered_small
    ### Read Excel Database
    df_baza = pd.read_excel(database_location, sheet_name='Baza de Date', dtype='object', keep_default_na=False, engine='openpyxl')
    
    week_first_day = datetime.strptime(f'{data_azi.isocalendar().year} {data_azi.isocalendar().week} 1', '%G %V %u')
    week_last_day = datetime.strptime(f'{data_azi.isocalendar().year} {data_azi.isocalendar().week} 7', '%G %V %u')
    
    df_baza['Data inregistrare Distrigaz'].replace(to_replace=r'-|Anulat|Fara|Alt Operator', value=np.nan, regex=True, inplace=True)
    df_baza['Data inregistrare Distrigaz'].replace(to_replace='(\d{2}).(\d{2}).(\d{4})', value='\\3-\\2-\\1 00:00:00', inplace=True, regex=True)
    df_baza['Data inregistrare Distrigaz'] = pd.to_datetime(df_baza['Data inregistrare Distrigaz'], format='%Y-%m-%d')
    df_baza['Data T1'] = df_baza['Data inregistrare Distrigaz'] + pd.offsets.DateOffset(years=2) - pd.offsets.DateOffset(weeks=number_of_weeks_T1)
    df_baza['Data expirare'] = df_baza['Data inregistrare Distrigaz'] + pd.offsets.DateOffset(years=2)

    ### Create df_filtered by date
    df_filtered = (df_baza['Data T1'] > str(week_first_day)) & (df_baza['Data T1'] <= str(week_last_day))
    df_filtered = df_baza.loc[df_filtered]
    df_filtered.sort_values(by=['Data T1', 'Numar inregistrare Distrigaz'], inplace=True)
    try:
        df_filtered['Data incheierii contractului'] = pd.to_datetime(df_filtered['Data incheierii contractului'], format='%Y-%m-%d')
    except:
        pass
    try:
        df_filtered['Data finalizarii contractului'] = pd.to_datetime(df_filtered['Data finalizarii contractului'], format='%Y-%m-%d')
    except:
        pass
    df_filtered['Data incheierii contractului'] = df_filtered['Data incheierii contractului'].dt.strftime('%d.%m.%Y')
    df_filtered['Data finalizarii contractului'] = df_filtered['Data finalizarii contractului'].dt.strftime('%d.%m.%Y')
    df_filtered['Data inregistrare Distrigaz'] = df_filtered['Data inregistrare Distrigaz'].dt.strftime('%d.%m.%Y')
    df_filtered['Data T1'] = df_filtered['Data T1'].dt.strftime('%d.%m.%Y')
    df_filtered['Data expirare'] = df_filtered['Data expirare'].dt.strftime('%d.%m.%Y')

    ### Set Table Columns to Show
    df_filtered_small = df_filtered[[
        'ID',
        'Data expirare',
        'SMS 3S DA/NU',
        'EMAIL 3S DA/NU',
        'SMS 5S DA/NU',
        'EMAIL 5S DA/NU',
        'Telefon',
        'Email',
        'PROGRAMAT DA/NU',
        'Persoana contact',
        'Adresa client',
        'ID Client',
        'Data inregistrare Distrigaz',
        'Numar contract',
        'Numar inregistrare Distrigaz',
        'Data incheierii contractului',
        'Data finalizarii contractului',
        'Tip proprietar',
        'Tip verificare / revizie'
        ]]

    row_count = len(df_filtered_small)
    column_name_list = df_filtered_small.columns.values.tolist()

try:
    ### Change the current working Directory
    os.chdir(os.path.join(os.getcwd(), '__Anonimized for GitHub\MCapp'))
except OSError:
    print("Can't change the Current Working Directory")
print("Current Working Directory ", os.getcwd())

### database_location and other variables
database_location = 'Test Database\\testing_database.xlsx'
telefon_programari = '0777.111.222 | 0722.22.222'
email_firma = 'contact@example-site.ro'
link_firma_programari = 'https://example-site.ro/programari/'

### Load JSONs
with open('JSON_samples\\judet_sat_dict_sample.json','r') as f:
    judet_sat_dict = json.load(f)
with open('JSON_samples\\streets_dict_sample.json','r') as f:
    streets_dict = json.load(f)
with open('JSON_samples\\nume_instalator_sample.json', 'r') as f:
    nume_instalator = json.load(f)
with open('JSON_samples\\tip_verificare_revizie_sample.json', 'r') as f:
    tip_verificare_revizie = json.load(f)
with open('JSON_samples\\tip_artera_sample.json', 'r') as f:
    tip_artera = json.load(f)
with open('JSON_samples\\tip_proprietar_sample.json', 'r') as f:
    tip_proprietar = json.load(f)
with open('JSON_samples\\weeks_T1_sample.json', 'r') as f:
    weeks_T1 = json.load(f)

number_of_weeks_T1 = 3
custom_title_label = ''
row_count = 0
column_name_list = ''
df_filtered = ''
df_filtered_small = ''
data_azi = get_today()

app = QApplication(sys.argv)
welcome = UserWelcomeWindow()
widget = QStackedWidget()
widget.addWidget(welcome)
widget.setFixedHeight(600)
widget.setFixedWidth(800)
widget.setWindowTitle('MC App')
widget.setWindowIcon(QtGui.QIcon('UIs\\MCapp_icon.png'))
widget.show()

try:
    sys.exit(app.exec())
except:
    print('Exiting')
